import os
import tkinter as tk
# from tkinter import ttk
import ttkbootstrap as ttk
from functions import successful
import sqlite3
# from tkcalendar import DateEntry
from openpyxl import load_workbook
import openpyxl
import modules.settings as setting
import modules.connect as connect
# import modules.reports_table as report_table

conn1 = connect.MainDB("X")


def back_to_pg(frame):
    for fr in frame.winfo_children():
        fr.destroy()
    report_section(frame, frame.winfo_screenheight(),
                   frame.winfo_screenwidth())


def print_to_printer(txt, dt, date, template, product="", chrg=None, type=""):
    report = load_workbook(
        filename=f"./templates/{template}.xlsx")
    entervalue = report.active
    # print(dt)
    if txt == "rawmaterials":
        entervalue["C5"] = date[0]
        entervalue["F5"] = date[1]
    elif txt == "labour":
        entervalue['C5'] = date[0]
        entervalue['F5'] = date[1]
        entervalue['G42'] = chrg[0]
        entervalue['G45'] = chrg[1]
    elif txt == "transport":
        entervalue['C5'] = date[0]
        entervalue['F5'] = date[1]
    elif txt == "production":
        entervalue['A1'] = f"{product.upper()} PRODUCTION REPORT"
        entervalue['D5'] = date[0]
        entervalue['H5'] = date[1]
        entervalue['E7'] = f"{product.upper()}"
        entervalue['F7'] = f"TOTAL {product.upper()}"
        entervalue['G7'] = f"FAILED {product.upper()}"
        entervalue['H7'] = f"FINAL {product.upper()}"
    else:
        entervalue['G5'] = date[0]
        entervalue['M5'] = date[1]

    for i in range(len(dt)):
        for j in range(len(dt[i])):
            entervalue.cell(row=i+8, column=j+1, value=dt[i][j])
    
    
    if type == "Credit":
        report.save(
        filename=f'''reports/{txt}/salesCreditor/'{date[0].replace("/","-")}'to'{date[1].replace("/","-")}'.xlsx''')
        successful( 
            f'''reports\\{txt}\\salesCreditor\\'{date[0].replace("/","-")}'to'{date[1].replace("/","-")}'.xlsx''')    
        
    elif product == "Brick":
        report.save(
            filename=f'''reports/{txt}/Brick/'{date[0].replace("/","-")}'to'{date[1].replace("/","-")}'.xlsx''')
        successful(
            f'''reports\\{txt}\\Brick\\'{date[0].replace("/","-")}'to'{date[1].replace("/","-")}'.xlsx''')
        
    elif product == "Block":
        report.save(
            filename=f'''reports/{txt}/Block/'{date[0].replace("/","-")}'to'{date[1].replace("/","-")}'.xlsx''')
        successful(
            f'''reports\\{txt}\\Block\\'{date[0].replace("/","-")}'to'{date[1].replace("/","-")}'.xlsx''')
        
    else:
        report.save(
            filename=f'''reports/{txt}/'{date[0].replace("/","-")}'to'{date[1].replace("/","-")}'.xlsx''')
        successful(
            f'''reports\\{txt}\\'{date[0].replace("/","-")}'to'{date[1].replace("/","-")}'.xlsx''')


def report_section(main_frame, Width, Height):

    def column(dt, w):
        coldata = {}
        col = []
        for i in range(len(dt)):
            coldata = {"text": dt[i], "stretch": True,
                       "anchor": "center", "width": w[i]}
            col.append(coldata)
        return col

    def report(frame, template, t, product, dt=[], col=None, w=None, txt="", date=(), f_size=9):
        delete_pages2(frame)
        rawmaterial_frame = ttk.Frame(frame)
        ttk.Label(rawmaterial_frame, bootstyle="dark", text=txt, font=("Open Sans", 20, "bold")).grid(
            row=0, column=0, columnspan=1, padx=10, pady=10)
        date_section = ttk.LabelFrame(
            rawmaterial_frame, bootstyle="primary", text='Date')
        ttk.Label(date_section, font=("Open Sans", 13), text="Start Date :").grid(
            row=0, column=0, padx=20, pady=20)
        ttk.Label(date_section, font=("Open Sans", 13), text=f"{date[0]}").grid(
            row=0, column=1, padx=5, pady=5, sticky="W")
        ttk.Label(date_section, font=("Open Sans", 13), text="End Date :").grid(
            row=0, column=2, padx=20, pady=20)
        ttk.Label(date_section, font=("Open Sans", 13), text=f"{date[1]}").grid(
            row=0, column=3, padx=5, pady=5, sticky="W")
        rawmaterial_report = ttk.LabelFrame(
            rawmaterial_frame, bootstyle="primary", text='Report')
        setting.Table(rawmaterial_report, column(
            col, w), dt, (0, 0), 10, 10, 4, f_size)

        operation_section = ttk.LabelFrame(
            rawmaterial_frame, bootstyle="primary", text='Operation')
        if txt == "Labour Report":
            labour_total_cost = sum(conn1.get_lab_out(date)[:-1])
            # print(conn1.get_lab_out(date)[:-1])
            # print(labour_total_cost)
            payment = conn1.get_lab_out(date)[-1]
            outstanding_amount = payment-labour_total_cost
            # print(outstanding_amount)
            lt = tk.StringVar(value=labour_total_cost)
            oa = tk.StringVar(value=outstanding_amount)
            lt.set(str(labour_total_cost))
            oa.set(str(outstanding_amount))
            ttk.Label(rawmaterial_report, font=("Open Sans", 13), text="Total Labour Cost:").grid(
                row=1, column=0, padx=5, pady=5, sticky="E")
            # setting.ContentEntry(rawmaterial_report, lt,
            #                      (1, 1), f_size=9).entry.config(state="readonly")
            ttk.Label(rawmaterial_report, font=("Open Sans", 13), text=f"{labour_total_cost:.2f}").grid(
                row=1, column=1, padx=20, pady=10)
            ttk.Label(rawmaterial_report, font=("Open Sans", 13), text=f"Outstanding Amount:").grid(
                row=2, column=0, padx=20, pady=10, sticky="E")
            # setting.ContentEntry(rawmaterial_report, oa,
            #                      (2, 1), f_size=9).entry.config(state="readonly")
            ttk.Label(rawmaterial_report, font=("Open Sans", 13), text=f"{outstanding_amount:.2f}").grid(
                row=2, column=1, padx=5, pady=5)

        ttk.Button(operation_section, text='Back', command=lambda: back_to_pg(
            frame), bootstyle="warning").grid(row=0, column=0, padx=10, pady=10)

        if txt == "Labour Report":
            ttk.Button(operation_section, text='Print Report', command=lambda: print_to_printer(t, dt, date, template, product, (labour_total_cost, outstanding_amount)), bootstyle="Danger").grid(
                row=0, column=1, padx=10, pady=10)
        else:
            ttk.Button(operation_section, text='Print Report', command=lambda: print_to_printer(t, dt, date, template, product), bootstyle="Danger").grid(
                row=0, column=1, padx=10, pady=10)
        date_section.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        if txt == "Sales Report":
            # print(date,product)
            dat = conn1.get_sales_credit(date, product)
            # print(dat)
            ttk.Button(operation_section, text="Print Creditors Report", command=lambda: print_to_printer(
                t, dat, date, "salesreportcredit", product, type="Credit"), bootstyle='success').grid(row=0, column=2, padx=10, pady=10)
            operation_section.grid(
                row=1, column=1, sticky="nsew", padx=10, pady=10)
            rawmaterial_report.grid(
                row=2, column=0, sticky="nsew", columnspan=2, padx=10, pady=10)
        else:
            operation_section.grid(
                row=1, column=1, sticky="nsew", padx=10, pady=10)
            rawmaterial_report.grid(
                row=2, column=0, sticky="nsew", columnspan=2, padx=10, pady=10)

        # tk.Grid.rowconfigure(date_section, 1, weight=1)
        # tk.Grid.columnconfigure(date_section, 0, weight=1)
        # tk.Grid.rowconfigure(rawmaterial_report, 2, weight=1)
        # tk.Grid.columnconfigure(rawmaterial_report, 0, weight=1)
        # tk.Grid.rowconfigure(operation_section, 3, weight=1)
        # tk.Grid.columnconfigure(operation_section, 0, weight=1)
        rawmaterial_frame.pack(fill="both")

    def hide_menu_indicator():
        rawmaterial_indicate.config(bg='#c3c3c3')
        labour_indicate.config(bg='#c3c3c3')
        selling_indicate.config(bg='#c3c3c3')
        production_indicate.config(bg="#c3c3c3")
        transport_indicate.config(bg="#c3c3c3")

    def delete_pages2(frame):
        for frame in frame.winfo_children():
            frame.destroy()

    def delete_pages():
        for frame in content_section.winfo_children():
            frame.destroy()

    def menu_indicate(lb, page):
        hide_menu_indicator()
        lb.config(bg='#158aff')
        delete_pages()
        page()

    def rawmaterial_report_page():
        def getvalue():
            rawmateriallist = (others, Flyash, Dust, Cement, Admixture)
            rawmaterialperlst = (othersper, flyashper,
                                 dustper, cementper, admixtureper)
            # print(conn1.get_rawMaterial_stock())
            for i in range(len(rawmateriallist)):
                if rawmateriallist[i].get() == "":
                    rawmateriallist[i].set("0")
                    rawmaterialperlst[i].set("0")

            total = 0
            result = conn1.get_rawMaterial_stock(
                date=(entrystartdate.entry.get(), entryenddate.entry.get()))
            for i, r in enumerate(result):
                if r in lst1:
                    rawmateriallist[i].set(result[r][0])
                    total += float(result[r][0])
                else:
                    continue
            # print(total)
            for i in range(len(rawmateriallist)):
                try:
                    rawmaterialperlst[i].set(
                        f'''{round((float(rawmateriallist[i].get()) / total)*100,2)} %''')
                except:
                    pass
            rawmateriallist2 = (others.get(), Flyash.get(),
                                Dust.get(), Cement.get(), Admixture.get())

            rawmaterialperlst2 = (othersper.get(), flyashper.get(),
                                  dustper.get(), cementper.get(), admixtureper.get())
            rawmaterial = ("Others", "Flyash", "Dust", "Cement", "Admixutre")

            rawmaterialdata = zip(
                rawmaterial, rawmateriallist2, rawmaterialperlst2)
            rawmaterial_col = ("Raw Materials", "Amount (in tonn)",
                               "PERCENTAGE(%)")
            w = (300, 300, 300)
            rawmaterialdata = list(rawmaterialdata)
            col = ("ID", "Purchase Date", "Party Name", "Vehicle No.",
                   "Raw Material", "Qty(in Tonn)/Bag", "Price", "Total")
            dt = conn1.get_raw_material_report(
                (entrystartdate.entry.get(), entryenddate.entry.get()))
            w = (40, 150, 120, 150, 150, 150, 130, 100)
            getdata_btn = ttk.Button(
                calculate_section, text='Show Report Party Wise', command=lambda: report(main_frame, "rawmaterialreport", "rawmaterials", "", dt, col, w, "Raw Material Report", (entrystartdate.entry.get(), entryenddate.entry.get()), 12), bootstyle="warning")
            getdata_btn.grid(row=7, column=0, padx=10, pady=10)

        def printvalue():
            rawmateriallist = (others, Flyash, Dust, Cement, Admixture)
            rawmaterialperlst = (othersper, flyashper,
                                 dustper, cementper, admixtureper)
            rawmaterial_report = load_workbook(
                filename="./templates/rawmaterialreport.xlsx")
            entervalue = rawmaterial_report.active
            entervalue["B6"] = (entrystartdate.entry.get())
            entervalue["D6"] = (entryenddate.entry.get())

            for i in range(10, len(rawmateriallist)+10):
                entervalue.cell(row=i, column=2, value=int(
                    rawmateriallist[i-10].get()))
                entervalue.cell(row=i, column=3, value=float(
                    rawmaterialperlst[i-10].get()[:-2]))

            rawmaterial_report.save(
                filename=f'''reports/rawmaterials/'{entrystartdate.entry.get().replace("/","-")}'to'{entryenddate.entry.get().replace("/","-")}'.xlsx''')
            successful(
                f'''reports\\rawmaterials\\'{entrystartdate.entry.get().replace("/","-")}'to'{entryenddate.entry.get().replace("/","-")}'.xlsx''')
            # os.startfile(".\\reports\\rawmaterials\\'"+str(entrystartdate.entry.get())+"'to'"+str(entryenddate.entry.get())+"'.xlsx")
            # os.open(f"../reports/rawmaterials/'{entrystartdate.entry.get()}'to'{entryenddate.entry.get()}'.xlsx")

    ############################################## rawmaterial frame ###############################################################

        rawmaterial_frame = ttk.Frame(content_section)

        lb = ttk.Label(rawmaterial_frame, text='Stock Details',
                       font=('Open Sans', 20, "bold"))

        lb.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        date_section = ttk.LabelFrame(rawmaterial_frame, text='Date Selection')
        date_section.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        calculate_section = ttk.LabelFrame(
            rawmaterial_frame, text='Raw Material Purchased')
        calculate_section.grid(row=2, column=0, padx=10,
                               pady=10, sticky="nsew")

    ############################################# date section ###########################################################

        setting.ContentLabel(date_section, "Start Date", (0, 0))
        setting.ContentLabel(date_section, "End Date", (0, 2))

        getdata_btn = ttk.Button(
            date_section, text='Show', command=getvalue, bootstyle="danger")

        entrystartdate = ttk.DateEntry(
            date_section, width=20, dateformat=r"%d/%m/%Y", bootstyle="primary")
        entryenddate = ttk.DateEntry(
            date_section, width=20, dateformat=r"%d/%m/%Y", bootstyle="primary")

        entrystartdate.grid(row=0, column=1, padx=5, pady=5)
        entryenddate.grid(row=0, column=3, padx=5, pady=5)
        getdata_btn.grid(row=1, column=0, padx=10, pady=10)

    #################################################### calculate section ################################################

        setting.ContentLabel(calculate_section, "Amount (in tonn)", (0, 1))
        setting.ContentLabel(calculate_section, "Percentage (in %)", (0, 2))

        lst1 = ("Others", "FlyAsh", "Dust", "Cement", "Admixture")

        for i in range(len(lst1)):
            setting.ContentLabel(calculate_section, lst1[i], (i+1, 0))

        Flyash = tk.StringVar()
        Dust = tk.StringVar()
        Cement = tk.StringVar()
        Admixture = tk.StringVar()
        others = tk.StringVar()
        lst2 = (others, Flyash, Dust, Cement, Admixture)
        for i in range(len(lst2)):
            setting.ContentEntry(
                calculate_section, lst2[i], (i+1, 1)).entry.config(state="readonly")

        flyashper = tk.StringVar()
        dustper = tk.StringVar()
        cementper = tk.StringVar()
        admixtureper = tk.StringVar()
        othersper = tk.StringVar()
        lst3 = (othersper, flyashper, dustper, cementper, admixtureper)
        for i in range(len(lst3)):
            setting.ContentEntry(
                calculate_section, lst3[i], (i+1, 2)).entry.config(state="readonly")

        rawmaterial_frame.pack()

    # End

    def labour_report_page():
        def getdata():
            startdate = entrystartdate.entry.get()
            enddate = entryenddate.entry.get()
            dt = conn1.get_labour_report((startdate, enddate))
            # print(dt)
            col = ("Date", "Product", "Total Units", " Total Labour Charges", "Loading Charge",
                   "Unloading Charge", "Payment Type", "Payment Amount")
            w = (80, 80, 100, 180, 120, 140, 130, 150)
            report(main_frame, "labourreport", "labour", "", dt, col, w, "Labour Report",
                   (startdate, enddate), 10)

        ################################################ labour frame ####################################################

        labour_frame = ttk.Frame(content_section)
        lb = ttk.Label(labour_frame, text='Labour Details',
                       font=('Open Sans', 20, "bold"))

        lb.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

        lp_section = ttk.LabelFrame(
            labour_frame, text='Labour Production Report', bootstyle="primary")
        lp_section.grid(row=1, column=0, sticky="nsew", padx=20, pady=20)

        ##################################################### lp_section ##############################################
        setting.ContentLabel(lp_section, "Start Date", (0, 0))
        setting.ContentLabel(lp_section, "End Date", (0, 2))

        entrystartdate = ttk.DateEntry(
            lp_section, width=20, dateformat=r"%d/%m/%Y", bootstyle="primary")
        entryenddate = ttk.DateEntry(
            lp_section, width=20, dateformat=r"%d/%m/%Y", bootstyle="primary")

        entrystartdate.grid(row=0, column=1, padx=5, pady=5)
        entryenddate.grid(row=0, column=3, padx=5, pady=5)

        getdata_btn = ttk.Button(
            lp_section, text='Show Labour Report', command=getdata, bootstyle="danger")
        getdata_btn.grid(row=2, column=2, columnspan=2, padx=10, pady=10)

        ################################################################# l_ul section ###########################################

        labour_frame.pack()
    # end

    def selling_report_page():
        s_totalbricks = tk.StringVar()
        s_totalblocks = tk.StringVar()

        def getdata():
            t_brick, t_block = conn1.get_total_sales(
                (entrystartdate.entry.get(), entryenddate.entry.get()))
            s_totalbricks.set(str(t_brick))
            s_totalblocks.set(str(t_block))

        def show():
            # getdata()
            col = ("ID", "Date", "Party", "Product", "Quantity", "Price", "Total", "Vehicle", "Number", "Transport Chrge",
                   "Loading", "Total Loading", "Unloading", "Total Unloading", "Total Charge", "Payment", "Remarks")
            w = (40, 60, 60, 60, 60, 60, 60, 60,
                 60, 60, 60, 60, 60, 60, 60, 60, 60)
            dt = conn1.get_sales_report(
                (entrystartdate.entry.get(), entryenddate.entry.get()), productvalue.get())
            report(main_frame, "salesreport", "sales", productvalue.get(), dt, col, w, "Sales Report", (entrystartdate.entry.get(
            ), entryenddate.entry.get()))
            # read()

        ####################################################### selling frame ###################################################

        selling_frame = ttk.Frame(content_section)
        date_section = ttk.LabelFrame(
            selling_frame, text='Date Selection', bootstyle="primary")
        total_sales_section = ttk.LabelFrame(
            selling_frame, text='Total Sales', bootstyle="primary")
        lb = ttk.Label(selling_frame, text='Sales Details',
                       font=('Open Sans', 20, "bold"))
        lb.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        date_section.grid(row=1, column=0, padx=20, pady=20, sticky="nsew")
        total_sales_section.grid(
            row=2, column=0, padx=20, pady=20, sticky="nsew")

        ########################################################### date Section ##################################################

        setting.ContentLabel(date_section, "Start Date", (0, 0))
        setting.ContentLabel(date_section, "End Date", (0, 2))

        entrystartdate = ttk.DateEntry(
            date_section, width=20, dateformat=r"%d/%m/%Y", bootstyle="primary")
        entryenddate = ttk.DateEntry(
            date_section, width=20, dateformat=r"%d/%m/%Y", bootstyle="primary")

        entrystartdate.grid(row=0, column=1, padx=5, pady=5)
        entryenddate.grid(row=0, column=3, padx=5, pady=5)

        setting.ContentLabel(date_section, "Select Product", (1, 0))

        productvalue = tk.StringVar()

        ttk.Radiobutton(date_section, text="Block", variable=productvalue, bootstyle="success-toolbutton",
                        value="Block").grid(row=1, column=1, padx=5, pady=5)
        ttk.Radiobutton(date_section, text="Brick", variable=productvalue, bootstyle="success-toolbutton",
                        value="Brick").grid(row=1, column=2, padx=5, pady=5)

        ttk.Button(date_section, text='Show Sales Report', command=show, bootstyle="danger").grid(
            row=2, column=2, columnspan=2, padx=10, pady=10)

        ######################################################### total sale section #########################################

        setting.ContentLabel(total_sales_section, "Brick Sold", (0, 0))
        setting.ContentLabel(total_sales_section, "Block Sold", (1, 0))

        setting.ContentEntry(total_sales_section, s_totalbricks,
                             (0, 1)).entry.config(state="readonly")
        setting.ContentEntry(total_sales_section, s_totalblocks,
                             (1, 1)).entry.config(state="readonly")
        ttk.Button(total_sales_section, text='Show Sales', command=getdata,
                   bootstyle="warning").grid(row=0, column=2, rowspan=2, padx=10, pady=10)

        selling_frame.pack()

    def production_report_page():
        def get_report():
            startdate = entrystartdate.entry.get()
            enddate = entryenddate.entry.get()
            rd = conn1.get_production_report(
                (startdate, enddate), productvalue.get())
            if productvalue.get() == "Brick":
                col = ("ID", "Date", "Recipe", "Patiya", "Bricks", "T. Bricks",
                       "Fail Bricks", "Final Bricks", "Total Labour Charge")
            else:
                col = ("ID", "Date", "Recipe", "Patiya", "Blocks", "T. Blocks",
                       "Fail Blocks", "Final Blocks", "Total Labour Charge")
            w = (40, 100, 100, 100, 100, 100, 110, 110, 200)
            report(main_frame, "productionreport", "production", productvalue.get(), rd, col, w, "Production Report",
                   (startdate, enddate), 12)
        ################################################ labour frame ####################################################

        labour_frame = ttk.Frame(content_section)
        lb = ttk.Label(labour_frame, text='Production Details',
                       font=('Open Sans', 20, "bold"))

        lb.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

        lp_section = ttk.LabelFrame(
            labour_frame, text='Production Report', bootstyle="primary")
        lp_section.grid(row=1, column=0, sticky="nsew", padx=20, pady=20)

        ##################################################### lp_section ##############################################
        setting.ContentLabel(lp_section, "Start Date", (0, 0))
        setting.ContentLabel(lp_section, "End Date", (0, 2))

        entrystartdate = ttk.DateEntry(
            lp_section, width=20, dateformat=r"%d/%m/%Y", bootstyle="primary")
        entryenddate = ttk.DateEntry(
            lp_section, width=20, dateformat=r"%d/%m/%Y", bootstyle="primary")

        entrystartdate.grid(row=0, column=1, padx=5, pady=5)
        entryenddate.grid(row=0, column=3, padx=5, pady=5)

        setting.ContentLabel(lp_section, "Select Product", (1, 0))

        productvalue = tk.StringVar()

        ttk.Radiobutton(lp_section, text="Block", variable=productvalue, bootstyle="success-toolbutton",
                        value="Block").grid(row=1, column=1, padx=5, pady=5)
        ttk.Radiobutton(lp_section, text="Brick", variable=productvalue, bootstyle="success-toolbutton",
                        value="Brick").grid(row=1, column=2, padx=5, pady=5)

        getdata_btn = ttk.Button(
            lp_section, text='Show Production Report', command=get_report, bootstyle="danger")
        getdata_btn.grid(row=2, column=2, columnspan=2, padx=10, pady=10)

        ################################################################# l_ul section ###########################################

        labour_frame.pack()

    def transport_report_page():
        def show():
            # getdata()
            col = ("ID", "Date", "Party", "Quantity", "Vehicle Number",
                   "Transport Chrge", "Remarks")
            w = (150, 150, 150, 140, 140, 140, 150)
            dt = conn1.getTransportReport(
                (entrystartdate.entry.get(), entryenddate.entry.get()), transporttype.get())
            report(main_frame, "transportreport", "transport", "", dt, col, w, "Transport Report", (entrystartdate.entry.get(
            ), entryenddate.entry.get()))
            # read()

        ####################################################### selling frame ###################################################

        transport_frame = ttk.Frame(content_section)
        date_section = ttk.LabelFrame(
            transport_frame, text='Date Selection', bootstyle="primary")
        lb = ttk.Label(transport_frame, text='Transport Details',
                       font=('Open Sans', 20, "bold"))
        lb.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        date_section.grid(row=1, column=0, padx=20, pady=20, sticky="nsew")

        ########################################################### date Section ##################################################

        setting.ContentLabel(date_section, "Start Date", (0, 0))
        setting.ContentLabel(date_section, "End Date", (0, 2))

        entrystartdate = ttk.DateEntry(
            date_section, width=20, dateformat=r"%d/%m/%Y", bootstyle="primary")
        entryenddate = ttk.DateEntry(
            date_section, width=20, dateformat=r"%d/%m/%Y", bootstyle="primary")

        entrystartdate.grid(row=0, column=1, padx=5, pady=5)
        entryenddate.grid(row=0, column=3, padx=5, pady=5)

        setting.ContentLabel(date_section, "Select Transport", (1, 0))

        transporttype = tk.StringVar()

        ttk.Radiobutton(date_section, text="Tractor", variable=transporttype, bootstyle="success-toolbutton",
                        value="Tractor").grid(row=1, column=1, padx=5, pady=5)
        ttk.Radiobutton(date_section, text="Dumper", variable=transporttype, bootstyle="success-toolbutton",
                        value="Dumper").grid(row=1, column=2, padx=5, pady=5)

        ttk.Button(date_section, text='Show Transport Report', command=show, bootstyle="danger").grid(
            row=2, column=2, columnspan=2, padx=10, pady=10)

        ######################################################### total sale section #########################################

        transport_frame.pack()

    ##################################################### main frame ###########################################################

    selling_frame = ttk.Frame(main_frame)
    selling_frame.pack(fill="both")
    tk.Grid.columnconfigure(selling_frame, 0, weight=1)
    tk.Grid.rowconfigure(selling_frame, 0, weight=1)
    tk.Grid.rowconfigure(selling_frame, 1, weight=0)

    content_section = ttk.Frame(selling_frame)
    menuSection = ttk.Frame(selling_frame, bootstyle="dark")

    ############################## Menu Frame #####################################################

    # Buttons
    setting.MainBtn(menuSection, txt="Raw Material", cmd=lambda: menu_indicate(
        rawmaterial_indicate, rawmaterial_report_page), grd=(0, 0))
    setting.MainBtn(menuSection, txt="Labour", cmd=lambda: menu_indicate(
        labour_indicate, labour_report_page), grd=(0, 1))
    setting.MainBtn(menuSection, txt="Sales", cmd=lambda: menu_indicate(
        selling_indicate, selling_report_page), grd=(0, 2))
    setting.MainBtn(menuSection, txt="Production", cmd=lambda: menu_indicate(
        production_indicate, production_report_page), grd=(0, 3))
    setting.MainBtn(menuSection, txt="Transport", cmd=lambda: menu_indicate(
        transport_indicate, transport_report_page), grd=(0, 4))

    # Labels
    rawmaterial_indicate = setting.MainLabel(menuSection, 120, (1, 0))
    labour_indicate = setting.MainLabel(menuSection, 120, (1, 1))
    selling_indicate = setting.MainLabel(menuSection, 120, (1, 2))
    production_indicate = setting.MainLabel(menuSection, 120, (1, 3))
    transport_indicate = setting.MainLabel(menuSection, 120, (1, 4))

    ################################### Pack section #############################################
    menuSection.configure(width=Width, height=Height)
    menuSection.pack(side="top", anchor="n", fill="x")
    content_section.configure(width=Width, height=Height)
    content_section.pack()
